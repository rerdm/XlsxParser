

from openpyxl import Workbook
import openpyxl
from DatePrefix import DatePrefix
import os
import sys
import string


class FilterAndCreateNewXlsx:
    def __init__(self,xlsx_file_to_parse):

        # This will be the new name for the parsed/shorted xlsx file
        self.new_name_of_document = xlsx_file_to_parse

        self.header_to_parse = "Product Groupe"  # Here you have to add the content from the column you want to filter
        self.filter_column_cells_for_string = "Bedroom"  # Here you have to add content what you want to filter

        # Load the file you want to parse
        self.excel_file = openpyxl.load_workbook(self.new_name_of_document)
        print(self.new_name_of_document)
        self.excel_sheet = self.excel_file['Tabelle1'] #'Sheet1']

        # Create a new workbook --> new filtered/shorted xlsx
        self.wb2 = Workbook()
        self.ws2 = self.wb2.active

        # Counter for count the lines in the new xlsx file
        self.lines = 1

        # Reading out the max rows and max columns of the input xlsx
        self.row_count = self.excel_sheet.max_row
        self.column_count_header = self.excel_sheet.max_column

        # This variable will save the cell coordinates form the filtered cells
        self.array_with_cell_coordinates_of_filtered_string = []

        self.header = []
        self.header_coordinates = []
        self.column_to_parse = ""

        # This sting method will generate a array with letters A to Z (uppercase)
        self.letters = list(string.ascii_uppercase)

        self.count_header = 0
        self.stop_column_header = self.letters[self.column_count_header - 1]
        self.end_of_header = self.stop_column_header + "1"


        print("Row_count                 :",self.row_count)
        print("Column_count_from_header  :",self.column_count_header)
        print("stopColumnHeader          :", self.stop_column_header)

        self.xls_line_in_list = []

    def parse_file(self):

        # Check the content of the header and append it to header array
        for column_of_cells in self.excel_sheet['A1':self.end_of_header]:
            for cell in column_of_cells:
                self.header_coordinates.append(cell.coordinate)

                if cell.value == self.header_to_parse:

                    self.column_to_parse = cell.coordinate
                    self.header.append(cell.value)
                    continue

                self.header.append(cell.value)
                self.count_header = self.count_header + 1


        for i in range(len(self.header)):

            if self.header[i] == self.header_to_parse:

                print("Column you want to filter for content : ",self.letters[i]," : Column to filter for:",self.filter_column_cells_for_string)
                self.aVerificationColumn = self.letters[i]

        print("Header : ",self.header)
        print("Header_coordinates : ",self.header_coordinates)

        self.checkIfTxtAlreadyExist(DatePrefix().get_prefix() + self.new_name_of_document)
        self.writexlsxHeader(self.header)

        for row_of_cells in self.excel_sheet[ self.column_to_parse: self.column_to_parse+str(self.row_count)]:

            for cell in row_of_cells:

                if cell.value == self.filter_column_cells_for_string:

                    self.array_with_cell_coordinates_of_filtered_string.append(cell.coordinate)

                    for l in range(len(self.header)):

                        self.xls_line_in_list.append(self.createNewArrayForXlsx(cell, self.letters[l]))

                    self.writeXlsxContentToNewXlsx(self.xls_line_in_list)
                    self.xls_line_in_list = []

        print("Array_with_cell_coordinates_of_filtered_string :", self.array_with_cell_coordinates_of_filtered_string)

    def createNewArrayForXlsx(self, cell, column):

         new_cords = str(cell.coordinate).replace(self.aVerificationColumn, column)

         return self.excel_sheet[new_cords].value

    def writexlsxHeader(self,header):

        for i in range(len(header)):
            coordinate = self.letters[i] + str(1)
            self.ws2[coordinate] = header[i]

    def writeXlsxContentToNewXlsx(self, array):

        for i in range(len(array)):

            coordinate = self.letters[i]+str(self.lines+1)
            self.ws2[coordinate] = array[i]

        self.lines = self.lines + 1
        self.wb2.save(DatePrefix().get_prefix() + self.new_name_of_document)


    def checkIfTxtAlreadyExist(self, txtFile):

        actualDir = os.listdir()

        for files in actualDir:
            if txtFile in actualDir:
                sys.exit(txtFile+" : Was already created - Stop the programm to prevent overriding")


if __name__ == '__main__':

    xlsx_files = []
    chosen_xlsx = 0
    xlsx_file_to_parse = []
    actualDir = os.listdir()

    for elements in actualDir:

        if elements.split(".")[-1] == "xlsx":
            xlsx_files.append(elements)

    if len(xlsx_files) == 0:

        sys.exit("No xlsx file in actual dir --> program end ")

    if len(xlsx_files) == 1:
        xlsx_file_to_parse = xlsx_files[0]

    if len(xlsx_files) > 1:

        print("""More than 1 xlsx file \nChoose xlsx file""")

        for i in range(len(xlsx_files)):
            print(i,"-",xlsx_files[i])

        chosen_xlsx = int(input("Wich file to parse : \n"))
        xlsx_file_to_parse = xlsx_files[chosen_xlsx]

    print("Selected file             :",xlsx_file_to_parse)

    new_xlsx = FilterAndCreateNewXlsx(xlsx_file_to_parse)
    new_xlsx.parse_file()

    #testghhtgjfggfhg




